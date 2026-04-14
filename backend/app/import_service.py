from __future__ import annotations

import csv
import io
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook

from .exam_service import normalize_answer, normalize_token


QUESTION_RE = re.compile(r"^(?:Q|S|SORU)?\s*(\d+)$", re.IGNORECASE)
STUDENT_ID_HEADERS = {"student_id", "ogrenci_no", "ogrencino", "student", "id", "numara"}
BOOKLET_HEADERS = {"booklet_code", "booklet", "kitapcik", "kitapcik_kodu", "grup", "group"}
SUMMARY_REPORT_HEADERS = {"dogru", "yanlis", "neti", "puani"}
CANONICAL_NO_HEADERS = {"canonical_no", "kanonik_no", "kanonik_soru_no"}
GROUP_LABEL_HEADERS = {"group_label", "grup", "grup_etiketi", "ders", "bolum"}
WEIGHT_HEADERS = {"weight", "agirlik", "puan", "soru_agirligi"}
FIXED_WIDTH_ANSWER_RE = re.compile(r"\s{2,}([A-E\s]{5,})\s*$", re.IGNORECASE)
FIXED_WIDTH_STUDENT_ID_RE = re.compile(r"\d{5,}")
FIXED_WIDTH_METADATA_RE = re.compile(r"(?P<sequence>\d{5})(?P<classroom>\d{2})(?P<exam_code>\d{4})(?P<student_suffix>\d{3})")
DOMINANT_FIXED_WIDTH_EXAM_CODE_RATIO = 0.8
BOOKLET_POSITION_SUFFIXES = ("position", "sira", "soru_sirasi")
BOOKLET_ANSWER_SUFFIXES = ("answer", "cevap", "dogru_cevap")
EXCEL_HEADER_SCORE_ALIASES = CANONICAL_NO_HEADERS | GROUP_LABEL_HEADERS | WEIGHT_HEADERS | BOOKLET_HEADERS | {"answer_key"}


def normalize_header(value: Any) -> str:
    raw = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-zA-Z0-9]+", "_", raw).strip("_").lower()


def parse_question_headers(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for raw_header in headers:
        match = QUESTION_RE.match(normalize_header(raw_header))
        if match:
            mapping[raw_header] = match.group(1)
    return mapping


def resolve_header(headers_by_normalized: dict[str, str], aliases: set[str]) -> str | None:
    for alias in aliases:
        if alias in headers_by_normalized:
            return headers_by_normalized[alias]
    return None


def parse_definition_booklet_columns(headers: list[str]) -> dict[str, dict[str, str]]:
    booklet_columns: dict[str, dict[str, str]] = {}
    for raw_header in headers:
        normalized = normalize_header(raw_header)
        for suffix in BOOKLET_POSITION_SUFFIXES:
            suffix_token = f"_{suffix}"
            if normalized.endswith(suffix_token):
                booklet_code = normalize_token(normalized[: -len(suffix_token)])
                if booklet_code:
                    booklet_columns.setdefault(booklet_code, {})["position"] = raw_header
                break
        for suffix in BOOKLET_ANSWER_SUFFIXES:
            suffix_token = f"_{suffix}"
            if normalized.endswith(suffix_token):
                booklet_code = normalize_token(normalized[: -len(suffix_token)])
                if booklet_code:
                    booklet_columns.setdefault(booklet_code, {})["answer"] = raw_header
                break
    return booklet_columns


def sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ";" if sample.count(";") > sample.count(",") else ","


def decode_bytes(blob: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1254", "latin-1"):
        try:
            return blob.decode(encoding)
        except UnicodeDecodeError:
            continue
    return blob.decode("utf-8", errors="ignore")


def is_fixed_width_candidate(text: str) -> bool:
    first_line = next((line.strip("\ufeff") for line in text.splitlines() if line.strip()), "")
    if not first_line:
        return False
    if any(delimiter in first_line for delimiter in (",", ";", "\t", "|")):
        return False
    return FIXED_WIDTH_ANSWER_RE.search(first_line) is not None


def read_text_rows(blob: bytes) -> list[dict[str, Any]]:
    decoded = decode_bytes(blob)
    reader = csv.DictReader(io.StringIO(decoded), delimiter=sniff_delimiter(decoded[:2048]))
    return list(reader)


def read_excel_rows(blob: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    best_headers: list[str] | None = None
    best_data_rows: list[tuple[Any, ...]] = []
    best_score = -1

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(item or "").strip() for item in rows[0]]
        normalized_headers = {normalize_header(header) for header in headers if str(header or "").strip()}
        score = sum(1 for header in normalized_headers if header in EXCEL_HEADER_SCORE_ALIASES)
        score += len(parse_question_headers(headers))
        score += sum(1 for mapping in parse_definition_booklet_columns(headers).values() if mapping.get("position"))

        if score > best_score:
            best_headers = headers
            best_data_rows = rows[1:]
            best_score = score

    if not best_headers:
        return []

    return [{best_headers[index]: raw_row[index] for index in range(len(best_headers))} for raw_row in best_data_rows]


def extract_fixed_width_metadata(prefix: str) -> dict[str, str]:
    match = FIXED_WIDTH_METADATA_RE.search(prefix)
    if not match:
        return {}

    metadata: dict[str, str] = {}
    classroom = match.group("classroom")
    exam_code = match.group("exam_code")
    if classroom and classroom != "00":
        metadata["classroom"] = classroom
    if exam_code and exam_code != "0000":
        metadata["exam_code"] = exam_code
    return metadata


def detect_dominant_fixed_width_exam_code(row_metadata: list[dict[str, str]]) -> str:
    exam_codes = [metadata.get("exam_code", "") for metadata in row_metadata if metadata.get("exam_code")]
    if len(exam_codes) < 2:
        return ""

    code, count = Counter(exam_codes).most_common(1)[0]
    return code if count / len(exam_codes) >= DOMINANT_FIXED_WIDTH_EXAM_CODE_RATIO else ""


def parse_fixed_width_vendor_rows(text: str, *, default_booklet_code: str = "") -> list[dict[str, Any]]:
    pending_rows: list[dict[str, Any]] = []
    for index, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.rstrip()
        if not line.strip():
            continue

        match = FIXED_WIDTH_ANSWER_RE.search(line)
        if not match:
            continue

        answer_blob = match.group(1).upper()
        if not answer_blob.strip():
            continue

        answers = {
            str(position): token
            for position, token in enumerate(answer_blob, start=1)
            if normalize_answer(token)
        }
        prefix = line[: match.start()]
        student_match = FIXED_WIDTH_STUDENT_ID_RE.search(prefix)
        pending_rows.append(
            {
                "student_id": student_match.group(0) if student_match else f"SATIR-{index}",
                "booklet_code": normalize_token(default_booklet_code),
                "answers": answers,
                "question_count": len(answer_blob),
                "source_row": index,
                "fixed_width_metadata": extract_fixed_width_metadata(prefix),
            }
        )

    dominant_exam_code = detect_dominant_fixed_width_exam_code(
        [row.get("fixed_width_metadata") or {} for row in pending_rows]
    )

    parsed_rows: list[dict[str, Any]] = []
    for row in pending_rows:
        metadata = row.pop("fixed_width_metadata", {})
        decoded_fields: dict[str, str] = {}
        classroom = metadata.get("classroom", "")
        if classroom:
            decoded_fields["classroom"] = classroom
        if dominant_exam_code:
            decoded_fields["exam_code"] = dominant_exam_code
        elif metadata.get("exam_code"):
            decoded_fields["exam_code"] = metadata["exam_code"]

        if decoded_fields:
            row["decoded_fields"] = decoded_fields
        parsed_rows.append(row)

    if not parsed_rows:
        raise ValueError("Header'siz sabit-genislik TXT icinde soru-cevap verisi bulunamadi.")
    return parsed_rows


def parse_tabular_student_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    headers = list(rows[0].keys())
    normalized_headers = {normalize_header(header): header for header in headers}
    question_headers = parse_question_headers(headers)
    if not question_headers:
        if SUMMARY_REPORT_HEADERS.issubset(set(normalized_headers.keys())):
            raise ValueError("Bu dosya soru-cevap exportu degil, ozet rapor. Import icin ogrenci cevap kolonlari gerekir.")
        raise ValueError("Soru kolonlari bulunamadi. Beklenen basliklar: Q1, Q2 veya S1, S2.")

    student_header = next((normalized_headers[key] for key in STUDENT_ID_HEADERS if key in normalized_headers), None)
    booklet_header = next((normalized_headers[key] for key in BOOKLET_HEADERS if key in normalized_headers), None)
    if not booklet_header:
        raise ValueError("Kitapcik kolonu bulunamadi. Beklenen basliklar: booklet_code, kitapcik veya grup.")

    parsed_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        parsed_rows.append(
            {
                "student_id": str(row.get(student_header) or f"SATIR-{index}").strip(),
                "booklet_code": normalize_token(row.get(booklet_header)),
                "answers": {position: normalize_answer(row.get(header)) for header, position in question_headers.items()},
                "source_row": index,
            }
        )
    return parsed_rows


def build_questions_from_mapping_rows(rows: list[dict[str, Any]], booklet_codes: list[str]) -> list[dict[str, Any]]:
    headers = list(rows[0].keys()) if rows else []
    normalized_headers = {normalize_header(header): header for header in headers}
    canonical_header = resolve_header(normalized_headers, CANONICAL_NO_HEADERS)
    group_header = resolve_header(normalized_headers, GROUP_LABEL_HEADERS)
    weight_header = resolve_header(normalized_headers, WEIGHT_HEADERS)
    if not canonical_header:
        raise ValueError("Detayli cevap anahtari dosyasinda canonical_no kolonu zorunludur.")

    questions: list[dict[str, Any]] = []
    for row in rows:
        canonical_no = int(row.get(canonical_header) or 0)
        if canonical_no <= 0:
            raise ValueError("Detayli cevap anahtarinda canonical_no pozitif olmalidir.")

        question = {
            "canonical_no": canonical_no,
            "group_label": str(row.get(group_header or "", "") or "Genel").strip() or "Genel",
            "weight": float(row.get(weight_header or "", 1) or 1),
            "booklet_mappings": {},
        }

        for booklet_code in booklet_codes:
            position_key = f"{normalize_header(booklet_code)}_position"
            answer_key = f"{normalize_header(booklet_code)}_answer"
            if position_key not in normalized_headers or answer_key not in normalized_headers:
                raise ValueError(f"{booklet_code} kitapcigi icin position/answer kolonlari eksik.")

            question["booklet_mappings"][booklet_code] = {
                "position": int(row.get(normalized_headers[position_key]) or 0),
                "correct_answer": normalize_answer(row.get(normalized_headers[answer_key])),
            }

        questions.append(question)

    return questions


def row_has_definition_content(
    row: dict[str, Any],
    canonical_header: str | None,
    group_header: str | None,
    weight_header: str | None,
    booklet_columns: dict[str, dict[str, str]],
) -> bool:
    candidate_headers = [
        canonical_header or "",
        group_header or "",
        weight_header or "",
    ]
    for mapping in booklet_columns.values():
        candidate_headers.append(mapping.get("position", ""))
        candidate_headers.append(mapping.get("answer", ""))

    return any(str(row.get(header) or "").strip() for header in candidate_headers if header)


def build_questions_from_definition_rows(
    rows: list[dict[str, Any]],
    booklet_codes: list[str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    headers = list(rows[0].keys()) if rows else []
    normalized_headers = {normalize_header(header): header for header in headers}
    canonical_header = resolve_header(normalized_headers, CANONICAL_NO_HEADERS)
    group_header = resolve_header(normalized_headers, GROUP_LABEL_HEADERS)
    weight_header = resolve_header(normalized_headers, WEIGHT_HEADERS)
    if not canonical_header:
        raise ValueError("Excel tanim dosyasinda canonical_no kolonu zorunludur.")

    detected_booklet_columns = parse_definition_booklet_columns(headers)
    detected_booklet_codes = [code for code, mapping in detected_booklet_columns.items() if mapping.get("position")]
    if not detected_booklet_codes:
        detected_booklet_codes = [normalize_token(code) for code in booklet_codes if normalize_token(code)]
        detected_booklet_columns = {
            booklet_code: {
                "position": normalized_headers.get(f"{normalize_header(booklet_code)}_position", ""),
                "answer": normalized_headers.get(f"{normalize_header(booklet_code)}_answer", ""),
            }
            for booklet_code in detected_booklet_codes
        }

    if not detected_booklet_codes:
        raise ValueError("Excel tanim dosyasinda en az bir kitapcik icin *_sira veya *_position kolonu bulunmalidir.")

    processed_rows = [
        row for row in rows if row_has_definition_content(row, canonical_header, group_header, weight_header, detected_booklet_columns)
    ]
    if not processed_rows:
        raise ValueError("Excel tanim dosyasinda islenecek dolu soru satiri bulunamadi.")

    questions: list[dict[str, Any]] = []
    seen_canonical: set[int] = set()
    weight_source = "explicit"
    any_answer_present = False
    all_answers_complete = True

    for row_index, row in enumerate(processed_rows, start=2):
        canonical_raw = row.get(canonical_header)
        if str(canonical_raw or "").strip() == "":
            raise ValueError(f"Excel taniminda {row_index}. satirda canonical_no bos birakilamaz.")

        try:
            canonical_no = int(canonical_raw)
        except (TypeError, ValueError) as error:
            raise ValueError(f"Excel taniminda {row_index}. satirdaki canonical_no sayi olmalidir.") from error

        if canonical_no <= 0:
            raise ValueError("Excel taniminda canonical_no pozitif olmalidir.")
        if canonical_no in seen_canonical:
            raise ValueError(f"Excel taniminda ayni canonical_no iki kez kullanildi: {canonical_no}")
        seen_canonical.add(canonical_no)

        raw_weight = row.get(weight_header or "", "")
        if str(raw_weight or "").strip() == "":
            weight_value = 1.0
            weight_source = "defaulted"
        else:
            try:
                weight_value = float(raw_weight)
            except (TypeError, ValueError) as error:
                raise ValueError(f"Excel taniminda {canonical_no}. sorunun agirligi sayi olmalidir.") from error
        if weight_value <= 0:
            raise ValueError(f"Excel taniminda {canonical_no}. soru icin agirlik sifirdan buyuk olmalidir.")

        question = {
            "canonical_no": canonical_no,
            "group_label": str(row.get(group_header or "", "") or "Genel").strip() or "Genel",
            "weight": weight_value,
            "booklet_mappings": {},
        }

        for booklet_code in detected_booklet_codes:
            column_mapping = detected_booklet_columns.get(booklet_code) or {}
            position_header = column_mapping.get("position", "")
            answer_header = column_mapping.get("answer", "")
            if not position_header:
                raise ValueError(f"Excel taniminda {booklet_code} kitapcigi icin position kolonu eksik.")

            raw_position = row.get(position_header)
            if str(raw_position or "").strip() == "":
                raise ValueError(f"Excel taniminda {canonical_no}. soru / {booklet_code} kitapcigi position bos birakilamaz.")
            try:
                position_value = int(raw_position)
            except (TypeError, ValueError) as error:
                raise ValueError(f"Excel taniminda {canonical_no}. soru / {booklet_code} kitapcigi position sayi olmalidir.") from error
            if position_value <= 0:
                raise ValueError(f"Excel taniminda {canonical_no}. soru / {booklet_code} kitapcigi position pozitif olmalidir.")

            correct_answer = normalize_answer(row.get(answer_header or "", ""))
            if correct_answer:
                any_answer_present = True
            else:
                all_answers_complete = False

            question["booklet_mappings"][booklet_code] = {
                "position": position_value,
                "correct_answer": correct_answer,
            }

        questions.append(question)

    questions.sort(key=lambda item: item["canonical_no"])
    return questions, {
        "booklet_codes": detected_booklet_codes,
        "question_count": len(questions),
        "canonical_mapping_source": "explicit",
        "weight_source": weight_source,
        "answer_source": "file" if all_answers_complete else ("partial-file" if any_answer_present else "missing"),
    }


def build_questions_from_sequential_rows(rows: list[dict[str, Any]], booklet_codes: list[str]) -> list[dict[str, Any]]:
    headers = list(rows[0].keys()) if rows else []
    normalized_headers = {normalize_header(header): header for header in headers}
    booklet_header = next((normalized_headers[key] for key in BOOKLET_HEADERS if key in normalized_headers), None)
    if not booklet_header:
        raise ValueError("Sirali cevap anahtarinda booklet_code kolonu zorunludur.")

    question_headers = parse_question_headers(headers)
    answer_key_header = normalized_headers.get("answer_key")
    if not question_headers and not answer_key_header:
        raise ValueError("Sirali cevap anahtarinda Q1...Qn kolonlari veya answer_key kolonu gerekir.")

    expected_booklets = [normalize_token(code) for code in booklet_codes if normalize_token(code)]
    booklet_answers: dict[str, list[str]] = {}
    for row in rows:
        booklet_code = normalize_token(row.get(booklet_header))
        if booklet_code not in expected_booklets:
            continue

        if question_headers:
            ordered_headers = sorted(question_headers.items(), key=lambda item: int(item[1]))
            answers = [normalize_answer(row.get(header)) for header, _ in ordered_headers]
        else:
            answers = [normalize_answer(character) for character in str(row.get(answer_key_header) or "").strip().upper()]

        booklet_answers[booklet_code] = answers

    missing_booklets = [code for code in expected_booklets if code not in booklet_answers]
    if missing_booklets:
        raise ValueError(f"Sirali cevap anahtarinda eksik kitapcik satiri var: {', '.join(missing_booklets)}")

    lengths = {booklet: len(answers) for booklet, answers in booklet_answers.items()}
    if len(set(lengths.values())) != 1:
        detail = ", ".join(f"{booklet}={count}" for booklet, count in sorted(lengths.items()))
        raise ValueError(f"Sirali cevap anahtarinda kitapcik soru sayilari uyusmuyor: {detail}")

    question_count = next(iter(lengths.values()), 0)
    questions: list[dict[str, Any]] = []
    for canonical_no in range(1, question_count + 1):
        booklet_mappings: dict[str, dict[str, Any]] = {}
        for booklet_code in expected_booklets:
            correct_answer = booklet_answers[booklet_code][canonical_no - 1]
            if not correct_answer:
                raise ValueError(f"Sirali cevap anahtarinda {booklet_code} / Q{canonical_no} bos birakilamaz.")
            booklet_mappings[booklet_code] = {"position": canonical_no, "correct_answer": correct_answer}
        questions.append(
            {
                "canonical_no": canonical_no,
                "group_label": "Genel",
                "weight": 1,
                "booklet_mappings": booklet_mappings,
            }
        )
    return questions


def build_questions_from_fixed_width_answer_key(rows: list[dict[str, Any]], booklet_code: str) -> list[dict[str, Any]]:
    if not rows:
        raise ValueError("Sabit-genislik cevap anahtarinda gecerli satir bulunamadi.")

    answers = rows[0].get("answers") or {}
    question_count = int(rows[0].get("question_count") or 0)
    questions: list[dict[str, Any]] = []
    for canonical_no in range(1, question_count + 1):
        correct_answer = normalize_answer(answers.get(str(canonical_no)))
        if not correct_answer:
            raise ValueError(f"Sabit-genislik cevap anahtarinda Q{canonical_no} bos birakilamaz.")
        questions.append(
            {
                "canonical_no": canonical_no,
                "group_label": "Genel",
                "weight": 1,
                "booklet_mappings": {
                    normalize_token(booklet_code): {"position": canonical_no, "correct_answer": correct_answer}
                },
            }
        )
    return questions


async def parse_upload_file(upload: UploadFile) -> tuple[list[dict[str, Any]], str]:
    blob = await upload.read()
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return parse_tabular_student_rows(read_excel_rows(blob)), "excel"

    decoded = decode_bytes(blob)
    if is_fixed_width_candidate(decoded):
        return parse_fixed_width_vendor_rows(decoded), "vendor-fixed"
    return parse_tabular_student_rows(read_text_rows(blob)), "text"


async def parse_answer_key_upload(upload: UploadFile, booklet_codes: list[str]) -> tuple[list[dict[str, Any]], str, str]:
    blob = await upload.read()
    suffix = Path(upload.filename or "").suffix.lower()
    normalized_booklets = [normalize_token(code) for code in booklet_codes if normalize_token(code)]

    if suffix in {".xlsx", ".xlsm"}:
        rows = read_excel_rows(blob)
    else:
        decoded = decode_bytes(blob)
        if is_fixed_width_candidate(decoded):
            if len(normalized_booklets) != 1:
                raise ValueError("Header'siz sabit-genislik cevap anahtari yalnizca tek kitapcikli sinavlarda kabul edilir.")
            fixed_rows = parse_fixed_width_vendor_rows(decoded, default_booklet_code=normalized_booklets[0])
            return (
                build_questions_from_fixed_width_answer_key(fixed_rows, normalized_booklets[0]),
                "answer-key-vendor-fixed",
                "sequential",
            )
        rows = read_text_rows(blob)

    if not rows:
        raise ValueError("Cevap anahtari dosyasinda islenecek satir bulunamadi.")

    headers = list(rows[0].keys())
    normalized_headers = {normalize_header(header): header for header in headers}
    if "canonical_no" in normalized_headers:
        return build_questions_from_mapping_rows(rows, normalized_booklets), "answer-key-mapping", "detailed"
    if any(key in normalized_headers for key in BOOKLET_HEADERS):
        return build_questions_from_sequential_rows(rows, normalized_booklets), "answer-key-sequential", "sequential"
    raise ValueError(
        "Cevap anahtari formati taninamadi. Detayli mapping icin canonical_no/A_position/A_answer; pratik format icin booklet_code ve Q1...Qn kullanin."
    )


async def parse_definition_upload(upload: UploadFile, booklet_codes: list[str]) -> tuple[list[dict[str, Any]], dict[str, Any], str]:
    blob = await upload.read()
    suffix = Path(upload.filename or "").suffix.lower()
    normalized_booklets = [normalize_token(code) for code in booklet_codes if normalize_token(code)]

    if suffix in {".xlsx", ".xlsm"}:
        rows = read_excel_rows(blob)
        import_format = "definition-excel"
    else:
        rows = read_text_rows(blob)
        import_format = "definition-text"

    if not rows:
        raise ValueError("Tanim dosyasinda islenecek satir bulunamadi.")

    questions, profile = build_questions_from_definition_rows(rows, normalized_booklets)
    return questions, profile, import_format
