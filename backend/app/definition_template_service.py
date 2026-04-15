from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFINITION_TEMPLATE_FILENAME = "akademisyen_sinav_tanim_sablonu_500.xlsx"
UNIVERSITY_TITLE = "İZMİR KÂTİP ÇELEBİ ÜNİVERSİTESİ"
TEMPLATE_TITLE = "Sınav Tanım Şablonu"
TITLE_FILL = PatternFill(fill_type="solid", fgColor="12355B")
SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F6690")
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F7FAFC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D7DEE8"),
    right=Side(style="thin", color="D7DEE8"),
    top=Side(style="thin", color="D7DEE8"),
    bottom=Side(style="thin", color="D7DEE8"),
)
HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
BODY_FONT = Font(name="Segoe UI", size=10, color="243B53")


def style_instruction_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A6"
    sheet.merge_cells("A1:G1")
    sheet["A1"] = UNIVERSITY_TITLE
    sheet["A1"].font = Font(name="Cambria", bold=True, size=18, color="FFFFFF")
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("A2:G2")
    sheet["A2"] = TEMPLATE_TITLE
    sheet["A2"].font = Font(name="Cambria", bold=True, size=13, color="12355B")
    sheet["A2"].fill = SUBTITLE_FILL
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("A4:G4")
    sheet["A4"] = "Kullanım ilkeleri"
    sheet["A4"].font = Font(name="Cambria", bold=True, size=12, color="12355B")
    sheet["A4"].fill = SUBTITLE_FILL

    instructions = [
        "1. Veri girişi için yalnız Sablon sayfasını kullanın; importer bu sayfanın ilk satırını kolon başlığı olarak bekler.",
        "2. Yalnız doldurduğunuz soru satırları işlenir. Boş bırakılan satırlar güvenle atlanır.",
        "3. Her kitapçık için *_sira ve *_cevap kolonlarını aynı satırda tutarlı doldurun.",
        "4. Türkçe başlık alias'ları desteklenir; gerekirse sağa yeni kitapçık kolonları ekleyebilirsiniz.",
        "5. Kurum dışı paylaşım öncesi dosya adını sınav kodu ve dönemiyle birlikte arşivleyin.",
    ]
    for row_index, text in enumerate(instructions, start=5):
        cell = sheet.cell(row=row_index, column=1, value=text)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        for column_index in range(1, 8):
            merged_cell = sheet.cell(row=row_index, column=column_index)
            merged_cell.border = THIN_BORDER
            if row_index % 2 == 1:
                merged_cell.fill = ALT_ROW_FILL

    sheet["A11"] = "Not"
    sheet["A11"].font = Font(name="Segoe UI", bold=True, size=11, color="12355B")
    sheet["A12"] = "Sablon sayfasında üstte ek açıklama satırı bırakmayın; başlık satırı ilk satırda kalmalıdır."
    sheet["A12"].font = BODY_FONT
    sheet["A12"].alignment = Alignment(wrap_text=True)
    sheet.merge_cells("A12:G12")

    widths = {"A": 26, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18}
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 22


def style_template_sheet(sheet) -> None:
    headers = [
        "kanonik_no",
        "grup_etiketi",
        "agirlik",
        "A_sira",
        "A_cevap",
        "B_sira",
        "B_cevap",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    sheet.sheet_view.showGridLines = True

    header_comments = {
        "A1": "Pozitif, benzersiz kanonik soru numarası girin.",
        "B1": "Ders, grup veya ölçme alanı etiketi.",
        "C1": "Soru puan ağırlığı. 3,33 veya 3.33 gibi ondalıklı değer kabul edilir.",
        "D1": "A kitapçığındaki fiziksel soru pozisyonu.",
        "E1": "A kitapçığındaki doğru seçenek.",
        "F1": "B kitapçığındaki fiziksel soru pozisyonu.",
        "G1": "B kitapçığındaki doğru seçenek.",
    }
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.coordinate in header_comments:
            cell.comment = Comment(header_comments[cell.coordinate], UNIVERSITY_TITLE)

    for row_index in range(2, 502):
        for column_index in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if column_index != 2 else "left", vertical="center")
            if row_index % 2 == 0:
                cell.fill = ALT_ROW_FILL

    column_widths = {
        "A": 14,
        "B": 24,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
    }
    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width
    sheet.row_dimensions[1].height = 24


def build_exam_definition_template_xlsx() -> bytes:
    workbook = Workbook()
    workbook.properties.creator = UNIVERSITY_TITLE
    workbook.properties.company = UNIVERSITY_TITLE
    workbook.properties.title = f"{UNIVERSITY_TITLE} {TEMPLATE_TITLE}"
    workbook.properties.subject = "Tam Excel tanımı için kurumsal sınav şablonu"

    instruction_sheet = workbook.active
    instruction_sheet.title = "Kullanim"
    style_instruction_sheet(instruction_sheet)

    template_sheet = workbook.create_sheet("Sablon")
    style_template_sheet(template_sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()