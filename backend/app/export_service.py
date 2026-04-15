from __future__ import annotations

import csv
from datetime import datetime
import html
import io
import json
from pathlib import Path
from typing import Any
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from odf.opendocument import OpenDocumentSpreadsheet
from odf.style import ParagraphProperties, Style, TableCellProperties, TextProperties
from odf.table import Table as OdfTable, TableCell, TableColumn, TableRow
from odf.text import P
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle

from .exam_service import enrich_session_for_reporting


WINDOWS_FONT_CANDIDATES = [
    {
        "regular": Path("C:/Windows/Fonts/calibri.ttf"),
        "bold": Path("C:/Windows/Fonts/calibrib.ttf"),
    },
    {
        "regular": Path("C:/Windows/Fonts/segoeui.ttf"),
        "bold": Path("C:/Windows/Fonts/segoeuib.ttf"),
    },
    {
        "regular": Path("C:/Windows/Fonts/arial.ttf"),
        "bold": Path("C:/Windows/Fonts/arialbd.ttf"),
    },
]
REPORT_TITLE = "İZMİR KÂTİP ÇELEBİ ÜNİVERSİTESİ SINAV SONUÇ RAPORU"
REPORT_ACCENT_FILL = PatternFill(fill_type="solid", fgColor="12355B")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E8FF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F6690")
BODY_ALT_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
WARNING_SECTION_FILL = PatternFill(fill_type="solid", fgColor="FEE4E2")
WARNING_HEADER_FILL = PatternFill(fill_type="solid", fgColor="B42318")
WARNING_BODY_FILL = PatternFill(fill_type="solid", fgColor="FEF3F2")
WARNING_TEXT_COLOR = "B42318"
THIN_BORDER = Border(
    left=Side(style="thin", color="D7DEE8"),
    right=Side(style="thin", color="D7DEE8"),
    top=Side(style="thin", color="D7DEE8"),
    bottom=Side(style="thin", color="D7DEE8"),
)
LTR_MARK = "\u200e"
LTR_SENSITIVE_COLUMNS = {"Sinif", "Form kodu", "Form tarihi"}
LTR_SENSITIVE_SUMMARY_LABELS = {"Sistem sinav kodu"}
WARNING_SCOPE_LABELS = {
    "summary_cards": "Genel özet",
    "booklet_table": "Kitapçık özeti",
    "group_table": "Grup özeti",
    "question_table": "Soru analizi",
    "question_choice_table": "Şık dağılımı",
    "student_table": "Öğrenci listesi",
}
WARNING_SCOPE_KEYS = {
    "missing_answer_layer": [],
    "inferred_canonical_mapping": ["question_table", "question_choice_table"],
    "defaulted_weights": ["summary_cards", "booklet_table", "group_table", "question_table", "question_choice_table", "student_table"],
}
REPORT_WARNING_COPY = {
    "missing_answer_layer": {
        "title": "Cevap anahtarı henüz tamamlanmadı",
        "message": "Tüm kitapçıkların doğru cevapları tamamlanmadan puanlama açılamaz.",
    },
    "inferred_canonical_mapping": {
        "title": "Kitapçık sırası ayrıca doğrulanmadı",
        "message": "Kitapçık sırası ayrı bir tanım kaynağından gelmediği için soru bazlı yorumları ön bilgi olarak değerlendirin.",
    },
    "defaulted_weights": {
        "title": "Soru puanları eşit kabul edildi",
        "message": "Soru ağırlıkları ayrıca tanımlanmadığı için puan ve yüzde hesapları eşit ağırlıkla yapıldı.",
    },
}


def resolve_session(exam: dict[str, Any], session_id: str) -> dict[str, Any]:
    sessions = list(exam.get("sessions", []))
    if not sessions:
        raise ValueError("Export icin once bir sonuc oturumu olusmalidir.")
    if session_id == "latest":
        return enrich_session_for_reporting(exam, sessions[0])
    for session in sessions:
        if session.get("session_id") == session_id:
            return enrich_session_for_reporting(exam, session)
    raise ValueError("Istenen sonuc oturumu bulunamadi.")


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    if isinstance(value, list):
        return " | ".join(stringify(item) for item in value)
    if isinstance(value, dict):
        return " | ".join(f"{key}: {stringify(item)}" for key, item in value.items())
    return str(value)


def format_export_datetime(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return text
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone()
    return parsed.strftime("%d.%m.%Y %H:%M")


def stringify_export_value(section: dict[str, Any], row: list[Any], column_index: int) -> str:
    value = row[column_index] if column_index < len(row) else ""
    text = stringify(value)
    if not text:
        return ""

    column_name = section["columns"][column_index] if column_index < len(section["columns"]) else ""
    summary_label = stringify(row[0]) if row else ""
    should_force_ltr = column_name in LTR_SENSITIVE_COLUMNS or (
        section.get("title") == "Genel Ozet"
        and column_name == "Deger"
        and summary_label in LTR_SENSITIVE_SUMMARY_LABELS
    )
    if should_force_ltr and not text.startswith(LTR_MARK):
        return f"{LTR_MARK}{text}"
    return text


def build_section(
    title: str,
    columns: list[str],
    rows: list[list[Any]],
    *,
    warning_columns: list[str] | None = None,
    warning_row_labels: list[str] | None = None,
    warning_row_indexes: list[int] | None = None,
    tone: str = "default",
) -> dict[str, Any]:
    return {
        "title": title,
        "columns": columns,
        "rows": rows,
        "warning_columns": list(dict.fromkeys(warning_columns or [])),
        "warning_row_labels": list(dict.fromkeys(warning_row_labels or [])),
        "warning_row_indexes": sorted(set(warning_row_indexes or [])),
        "tone": tone,
    }


def get_warning_row_indexes(section: dict[str, Any]) -> set[int]:
    warning_indexes = set(section.get("warning_row_indexes") or [])
    warning_labels = {str(value) for value in (section.get("warning_row_labels") or [])}
    if warning_labels:
        for row_index, row in enumerate(section.get("rows") or []):
            first_value = stringify(row[0] if row else "")
            if first_value in warning_labels:
                warning_indexes.add(row_index)
    return warning_indexes


def build_warning_scope_text(warning_code: str, analysis_integrity: dict[str, Any]) -> str:
    affected_columns = analysis_integrity.get("affected_columns") or {}
    scope_keys = WARNING_SCOPE_KEYS.get(warning_code)
    if scope_keys is None:
        scope_keys = [key for key, values in affected_columns.items() if values]

    scope_parts = []
    for scope_key in scope_keys:
        scoped_columns = [str(value) for value in (affected_columns.get(scope_key) or []) if str(value)]
        if not scoped_columns:
            continue
        scope_parts.append(f"{WARNING_SCOPE_LABELS.get(scope_key, scope_key)}: {', '.join(scoped_columns)}")

    if scope_parts:
        return "Bu raporda dikkat edilecek başlıklar: " + " / ".join(scope_parts)
    return "Bu uyarı, rapor okunurken ilgili adımın tamamlanması gerektiğini hatırlatır."


def build_preparation_warning_rows(session: dict[str, Any]) -> list[list[str]]:
    analysis_integrity = session.get("analysis_integrity") or {}
    warning_rows: list[list[str]] = []
    for warning in analysis_integrity.get("warnings") or []:
        warning_code = str(warning.get("code") or "")
        override = REPORT_WARNING_COPY.get(warning_code, {})
        title = str(override.get("title") or warning.get("title") or "Hazırlık notu")
        message = str(override.get("message") or warning.get("message") or "")
        scope = build_warning_scope_text(warning_code, analysis_integrity)
        warning_rows.append([title, message, scope])
    return warning_rows


def question_numbers_from_session(session: dict[str, Any]) -> list[int]:
    numbers = [int(item.get("canonical_no")) for item in (session.get("question_summary") or []) if item.get("canonical_no")]
    if numbers:
        return numbers
    first_student = next(iter(session.get("student_results") or []), None)
    if not first_student:
        return []
    return [int(item.get("canonical_no")) for item in first_student.get("question_responses") or [] if item.get("canonical_no")]


def response_lookup(student_row: dict[str, Any]) -> dict[int, str]:
    return {
        int(item.get("canonical_no")): stringify(item.get("display_value") or item.get("marked_answer") or "-(B)")
        for item in (student_row.get("question_responses") or [])
        if item.get("canonical_no") is not None
    }


def position_numbers_from_session(session: dict[str, Any]) -> list[int]:
    student_rows = list(session.get("student_results") or session.get("student_preview") or [])
    max_position = 0
    for student_row in student_rows:
        for item in student_row.get("question_responses") or []:
            max_position = max(max_position, int(item.get("booklet_position") or item.get("canonical_no") or 0))
    return list(range(1, max_position + 1))


def position_response_lookup(student_row: dict[str, Any]) -> dict[int, str]:
    return {
        int(item.get("booklet_position") or item.get("canonical_no")): stringify(item.get("display_value") or item.get("marked_answer") or "-(B)")
        for item in (student_row.get("question_responses") or [])
        if item.get("canonical_no") is not None
    }


def format_choice_distribution_cell(row: dict[str, Any], choice: str) -> str:
    distribution = (row.get("choice_distribution") or {}).get(choice) or {}
    count = distribution.get("count")
    rate = distribution.get("rate")
    if count in (None, "") and rate in (None, ""):
        return ""
    return f"{stringify(count)} / %{stringify(rate)}"


def build_methodology_rows(glossary: list[dict[str, Any]], commentary: list[dict[str, Any]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for item in commentary:
        rows.append([
            "Akademik yorum",
            str(item.get("title") or ""),
            str(item.get("text") or ""),
            "",
        ])
    for item in glossary:
        rows.append([
            "Metodoloji",
            str(item.get("term") or ""),
            str(item.get("how") or ""),
            " ".join(part for part in [str(item.get("why") or ""), str(item.get("value_use") or "")] if part).strip(),
        ])
    return rows


def build_report_sections(
    exam: dict[str, Any],
    session: dict[str, Any],
    *,
    response_chunk_size: int | None,
) -> list[dict[str, Any]]:
    summary = session.get("summary") or {}
    net_policy = session.get("net_policy") or {}
    highlights = session.get("assessment_highlights") or {}
    glossary = session.get("analysis_glossary") or []
    commentary = session.get("academic_commentary") or []
    student_rows = list(session.get("student_results") or session.get("student_preview") or [])
    analysis_integrity = session.get("analysis_integrity") or {}
    summary_warning_labels = [str(value) for value in (analysis_integrity.get("affected_columns") or {}).get("summary_cards", [])]
    booklet_warning_columns = [str(value) for value in (analysis_integrity.get("affected_columns") or {}).get("booklet_table", [])]
    group_warning_columns = [str(value) for value in (analysis_integrity.get("affected_columns") or {}).get("group_table", [])]
    question_warning_columns = [str(value) for value in (analysis_integrity.get("affected_columns") or {}).get("question_table", [])]
    question_choice_warning_columns = [str(value) for value in (analysis_integrity.get("affected_columns") or {}).get("question_choice_table", [])]
    student_warning_columns = [str(value) for value in (analysis_integrity.get("affected_columns") or {}).get("student_table", [])]
    preparation_warning_rows = build_preparation_warning_rows(session)

    sections: list[dict[str, Any]] = [
        build_section(
            "Genel Ozet",
            ["Alan", "Deger"],
            [
                ["Sistem sinav kodu", exam.get("exam_code", "")],
                ["Sinav basligi", exam.get("title", "")],
                ["Sinav yili", exam.get("exam_year", "")],
                ["Donem", exam.get("exam_term", "")],
                ["Sinav turu", exam.get("exam_type", "")],
                ["Olusturma", format_export_datetime(session.get("created_at", ""))],
                ["Import formati", session.get("import_format", "")],
                ["Net kurali", net_policy.get("label") or summary.get("net_policy_label") or "Belirtilmedi"],
                ["Ogrenci sayisi", summary.get("student_count", "")],
                ["Toplam soru", summary.get("total_questions", "")],
                ["Ortalama puan", summary.get("average_score", "")],
                ["Ortalama dogru", summary.get("average_correct_count", "")],
                ["Ortalama yanlis", summary.get("average_wrong_count", "")],
                ["Ortalama bos", summary.get("average_blank_count", "")],
                ["Ortalama net", summary.get("average_net_count", "")],
                ["Ort. yüzde", summary.get("average_percent", "")],
                ["Tamamlama orani", summary.get("completion_rate", "")],
                ["Guvenirlik alpha", summary.get("reliability_alpha", "")],
                ["Standart sapma", summary.get("score_std_dev", "")],
                ["Atlanan satir", summary.get("skipped_rows", "")],
            ],
            warning_row_labels=summary_warning_labels,
        ),
        build_section(
            "Hazırlık Bütünlüğü Uyarıları",
            ["Başlık", "Açıklama", "Bu raporda dikkat edilecek alanlar"],
            preparation_warning_rows,
            warning_row_indexes=list(range(len(preparation_warning_rows))),
            tone="warning",
        ) if preparation_warning_rows else build_section("", [], []),
        build_section(
            "Sinava Ozel Akademik Yorum",
            ["Baslik", "Yorum"],
            [[item.get("title", ""), item.get("text", "")] for item in commentary],
        ),
        build_section(
            "Rapor Metodolojisi ve Terim Aciklamalari",
            ["Tur", "Gosterge", "Uretim ozeti", "Kullanim ve yorum"],
            build_methodology_rows(glossary, commentary),
        ),
        build_section(
            "Olcme Degerlendirme Uyarilari",
            ["Baslik", "Icerik"],
            [
                [
                    "En zor sorular",
                    " | ".join(
                        f"S{item.get('canonical_no')} (%{stringify(item.get('correct_rate'))})"
                        for item in highlights.get("hardest_questions", [])
                    ),
                ],
                [
                    "En kolay sorular",
                    " | ".join(
                        f"S{item.get('canonical_no')} (%{stringify(item.get('correct_rate'))})"
                        for item in highlights.get("easiest_questions", [])
                    ),
                ],
                [
                    "En cok bos birakilan sorular",
                    " | ".join(
                        f"S{item.get('canonical_no')} (%{stringify(item.get('blank_rate'))})"
                        for item in highlights.get("most_blank_questions", [])
                    ),
                ],
                [
                    "Ayirt ediciligi zayif sorular",
                    " | ".join(
                        f"S{item.get('canonical_no')} ({stringify(item.get('discrimination_index'))})"
                        for item in highlights.get("weak_discrimination_questions", [])
                    ),
                ],
                [
                    "En yuksek performans",
                    " | ".join(
                        f"#{item.get('exam_rank')} {stringify(item.get('student_id'))} {stringify(item.get('student_full_name'))}"
                        for item in highlights.get("top_students", [])
                    ),
                ],
                [
                    "En dusuk performans",
                    " | ".join(
                        f"#{item.get('exam_rank')} {stringify(item.get('student_id'))} {stringify(item.get('student_full_name'))}"
                        for item in highlights.get("lowest_students", [])
                    ),
                ],
            ],
        ),
        build_section(
            "Kitapcik Ozeti",
            [
                "Kitapcik",
                "Ogrenci",
                "Toplam puan",
                "Toplam D",
                "Toplam Y",
                "Toplam B",
                "Toplam net",
                "Ort. puan",
                "Maks.",
                "Ort. D",
                "Ort. Y",
                "Ort. B",
                "Ort. net",
                "Ort. dogruluk",
            ],
            [
                [
                    row.get("booklet_code", ""),
                    row.get("student_count", ""),
                    row.get("total_score", ""),
                    row.get("total_correct_count", ""),
                    row.get("total_wrong_count", ""),
                    row.get("total_blank_count", ""),
                    row.get("total_net_count", ""),
                    row.get("average_score", ""),
                    row.get("max_score", ""),
                    row.get("average_correct_count", ""),
                    row.get("average_wrong_count", ""),
                    row.get("average_blank_count", ""),
                    row.get("average_net_count", ""),
                    row.get("average_accuracy_percent", ""),
                ]
                for row in (session.get("booklet_summary") or [])
            ],
            warning_columns=booklet_warning_columns,
        ),
        build_section(
            "Grup Ozeti",
            [
                "Grup",
                "Ogrenci",
                "Soru",
                "Toplam puan",
                "Toplam D",
                "Toplam Y",
                "Toplam B",
                "Toplam net",
                "Maks.",
                "Ort. puan",
                "Ort. D",
                "Ort. Y",
                "Ort. B",
                "Ort. net",
                "Ort. dogruluk",
            ],
            [
                [
                    row.get("group_label", ""),
                    row.get("student_count", ""),
                    row.get("question_count", ""),
                    row.get("total_score", ""),
                    row.get("total_correct_count", ""),
                    row.get("total_wrong_count", ""),
                    row.get("total_blank_count", ""),
                    row.get("total_net_count", ""),
                    row.get("max_score", ""),
                    row.get("average_score", ""),
                    row.get("average_correct_count", ""),
                    row.get("average_wrong_count", ""),
                    row.get("average_blank_count", ""),
                    row.get("average_net_count", ""),
                    row.get("average_accuracy_percent", ""),
                ]
                for row in (session.get("group_summary") or [])
            ],
            warning_columns=group_warning_columns,
        ),
        build_section(
            "Soru Bazlı Dağılım",
            [
                "Soru",
                "Grup",
                "Ağırlık",
                "Dogru %",
                "Yanlis %",
                "Bos %",
                "Gucluk indeksi",
                "Madde varyansi",
                "Madde std sapmasi",
                "Nokta-cift serili r",
                "r yorumu",
                "Ust grup %",
                "Alt grup %",
                "Ayirt indeksi",
                "Gucluk",
                "Ayirt edicilik",
                "Kitapçık sırası",
                "Anahtar",
            ],
            [
                [
                    f"S{row.get('canonical_no')}",
                    row.get("group_label", ""),
                    row.get("weight", ""),
                    row.get("correct_rate", ""),
                    row.get("wrong_rate", ""),
                    row.get("blank_rate", ""),
                    row.get("difficulty_index", ""),
                    row.get("item_variance", ""),
                    row.get("item_std_dev", ""),
                    row.get("point_biserial", ""),
                    row.get("point_biserial_label", ""),
                    row.get("upper_group_correct_rate", ""),
                    row.get("lower_group_correct_rate", ""),
                    row.get("discrimination_index", ""),
                    row.get("difficulty_label", ""),
                    row.get("discrimination_label", ""),
                    row.get("booklet_positions", {}),
                    row.get("booklet_answer_key", {}),
                ]
                for row in (session.get("question_summary") or [])
            ],
            warning_columns=question_warning_columns,
        ),
        build_section(
            "Sorulara Siklara Gore Verilen Cevaplarin Dagilimi",
            [
                "Soru",
                "Grup",
                "Ağırlık",
                "Ogrenci",
                "A",
                "B",
                "C",
                "D",
                "E",
                "Bos",
                "Dogru %",
                "Yanlis %",
                "Gucluk indeksi",
                "Madde varyansi",
                "Madde std sapmasi",
                "Nokta-cift serili r",
                "r yorumu",
                "Ust grup %",
                "Alt grup %",
                "Ayirt indeksi",
                "Gucluk",
                "Ayirt edicilik",
                "Kitapçık sırası",
                "Anahtar",
            ],
            [
                [
                    f"S{row.get('canonical_no')}",
                    row.get("group_label", ""),
                    row.get("weight", ""),
                    row.get("student_count", ""),
                    format_choice_distribution_cell(row, "A"),
                    format_choice_distribution_cell(row, "B"),
                    format_choice_distribution_cell(row, "C"),
                    format_choice_distribution_cell(row, "D"),
                    format_choice_distribution_cell(row, "E"),
                    f"{stringify((row.get('blank_distribution') or {}).get('count'))} / %{stringify((row.get('blank_distribution') or {}).get('rate'))}",
                    row.get("correct_rate", ""),
                    row.get("wrong_rate", ""),
                    row.get("difficulty_index", ""),
                    row.get("item_variance", ""),
                    row.get("item_std_dev", ""),
                    row.get("point_biserial", ""),
                    row.get("point_biserial_label", ""),
                    row.get("upper_group_correct_rate", ""),
                    row.get("lower_group_correct_rate", ""),
                    row.get("discrimination_index", ""),
                    row.get("difficulty_label", ""),
                    row.get("discrimination_label", ""),
                    row.get("booklet_positions", {}),
                    row.get("booklet_answer_key", {}),
                ]
                for row in (session.get("question_summary") or [])
            ],
            warning_columns=question_choice_warning_columns,
        ),
        build_section(
            "Ogrenci Onizlemi",
            [
                "Ogrenci no",
                "Ad",
                "Soyad",
                "Ad Soyad",
                "Sinif",
                "Kitapcik",
                "Form kodu",
                "Form tarihi",
                "D",
                "Y",
                "B",
                "Net",
                "Toplam",
                "Genel sira",
                "Sinif sira",
                "Puan",
                "Yüzde",
            ],
            [
                [
                    row.get("student_id", ""),
                    row.get("student_name", ""),
                    row.get("student_surname", ""),
                    row.get("student_full_name", ""),
                    row.get("classroom", ""),
                    row.get("booklet_code", ""),
                    row.get("scanned_exam_code", ""),
                    row.get("scanned_exam_date", ""),
                    row.get("correct_count", ""),
                    row.get("wrong_count", ""),
                    row.get("blank_count", ""),
                    row.get("net_count", ""),
                    row.get("total_questions", ""),
                    row.get("exam_rank", ""),
                    row.get("class_rank", ""),
                    row.get("score", ""),
                    row.get("weighted_percent", ""),
                ]
                for row in student_rows
            ],
            warning_columns=student_warning_columns,
        ),
    ]
    sections = [section for section in sections if section.get("columns")]

    question_numbers = question_numbers_from_session(session)
    position_numbers = position_numbers_from_session(session)
    if not question_numbers and not position_numbers:
        return sections

    if question_numbers:
        chunk_size = response_chunk_size or len(question_numbers)
        for start_index in range(0, len(question_numbers), chunk_size):
            question_chunk = question_numbers[start_index:start_index + chunk_size]
            columns = ["Genel sira", "Ogrenci no", "Ad Soyad", "Sinif", "Kitapcik"] + [f"S{number}" for number in question_chunk]
            rows: list[list[Any]] = []
            for student_row in student_rows:
                responses = response_lookup(student_row)
                rows.append(
                    [
                        student_row.get("exam_rank", ""),
                        student_row.get("student_id", ""),
                        student_row.get("student_full_name", ""),
                        student_row.get("classroom", ""),
                        student_row.get("booklet_code", ""),
                        *[responses.get(number, "-(B)") for number in question_chunk],
                    ]
                )
            sections.append(
                build_section(
                    f"Ogrenci Cevap Matrisi (Kanonik soru) S{question_chunk[0]}-S{question_chunk[-1]}",
                    columns,
                    rows,
                )
            )

    if position_numbers:
        chunk_size = response_chunk_size or len(position_numbers)
        for start_index in range(0, len(position_numbers), chunk_size):
            position_chunk = position_numbers[start_index:start_index + chunk_size]
            columns = ["Genel sira", "Ogrenci no", "Ad Soyad", "Sinif", "Kitapcik"] + [f"P{number}" for number in position_chunk]
            rows = []
            for student_row in student_rows:
                positions = position_response_lookup(student_row)
                rows.append(
                    [
                        student_row.get("exam_rank", ""),
                        student_row.get("student_id", ""),
                        student_row.get("student_full_name", ""),
                        student_row.get("classroom", ""),
                        student_row.get("booklet_code", ""),
                        *[positions.get(number, "-(B)") for number in position_chunk],
                    ]
                )
            sections.append(
                build_section(
                    f"Ogrenci Cevap Matrisi (Form pozisyonu) P{position_chunk[0]}-P{position_chunk[-1]}",
                    columns,
                    rows,
                )
            )

    return sections


def section_to_dicts(section: dict[str, Any]) -> list[dict[str, Any]]:
    columns = section["columns"]
    return [{column: row[index] if index < len(row) else "" for index, column in enumerate(columns)} for row in section["rows"]]


def build_csv_bytes(exam: dict[str, Any], session: dict[str, Any]) -> bytes:
    sections = build_report_sections(exam, session, response_chunk_size=30)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Rapor Basligi", REPORT_TITLE])
    writer.writerow([])
    for section in sections:
        writer.writerow(["Rapor Bolumu", section["title"]])
        writer.writerow(section["columns"])
        for row in section["rows"]:
            writer.writerow([stringify_export_value(section, row, index) for index in range(len(section["columns"]))])
        writer.writerow([])
    return buffer.getvalue().encode("utf-8-sig")


def build_json_bytes(exam: dict[str, Any], session: dict[str, Any]) -> bytes:
    sections = build_report_sections(exam, session, response_chunk_size=None)
    payload = {
        "report_title": REPORT_TITLE,
        "exam": {
            "exam_code": exam.get("exam_code"),
            "title": exam.get("title"),
            "exam_year": exam.get("exam_year", ""),
            "exam_term": exam.get("exam_term", ""),
            "exam_type": exam.get("exam_type", ""),
            "booklet_codes": exam.get("booklet_codes", []),
            "question_count": len(exam.get("questions", [])),
        },
        "session": session,
        "report_blocks": [
            {
                "title": section["title"],
                "columns": section["columns"],
                "warning_columns": section.get("warning_columns") or [],
                "warning_row_labels": section.get("warning_row_labels") or [],
                "warning_row_indexes": section.get("warning_row_indexes") or [],
                "tone": section.get("tone") or "default",
                "rows": section_to_dicts(section),
            }
            for section in sections
        ],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def render_text_table(section: dict[str, Any]) -> str:
    rows = [
        [stringify_export_value(section, row, index) for index in range(len(section["columns"]))]
        for row in section["rows"]
    ]
    headers = [stringify(value) for value in section["columns"]]
    widths = [len(header) for header in headers]
    for row in rows:
        for index, cell in enumerate(row):
            widths[index] = max(widths[index], len(cell))

    def render_row(values: list[str]) -> str:
        return " | ".join(value.ljust(widths[index]) for index, value in enumerate(values))

    separator = "-+-".join("-" * width for width in widths)
    lines = [section["title"], separator, render_row(headers), separator]
    for row in rows:
        lines.append(render_row(row))
    lines.append(separator)
    return "\n".join(lines)


def build_txt_bytes(exam: dict[str, Any], session: dict[str, Any]) -> bytes:
    sections = build_report_sections(exam, session, response_chunk_size=20)
    title_block = f"{REPORT_TITLE}\n{'=' * len(REPORT_TITLE)}"
    content = f"{title_block}\n\n" + "\n\n".join(render_text_table(section) for section in sections)
    return content.encode("utf-8")


def autosize_sheet(sheet) -> None:
    for column_index, column_cells in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=3, max_row=sheet.max_row), start=1):
        max_length = 0
        column_letter = get_column_letter(column_index)
        for cell in column_cells:
            max_length = max(max_length, len(stringify(cell.value)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def write_section_to_sheet(sheet, section: dict[str, Any]) -> None:
    column_count = max(len(section["columns"]), 1)
    warning_columns = set(section.get("warning_columns") or [])
    warning_row_indexes = get_warning_row_indexes(section)
    is_warning_section = section.get("tone") == "warning"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    sheet.cell(row=1, column=1, value=REPORT_TITLE)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(name="Cambria", bold=True, size=18, color="FFFFFF")
    title_cell.fill = REPORT_ACCENT_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    section_cell = sheet.cell(row=2, column=1, value=section["title"])
    section_cell.font = Font(name="Cambria", bold=True, size=12, color="12355B")
    section_cell.fill = WARNING_SECTION_FILL if is_warning_section else SECTION_FILL
    section_cell.alignment = Alignment(horizontal="left", vertical="center")

    sheet.append(section["columns"])
    for cell in sheet[3]:
        is_warning_header = stringify(cell.value) in warning_columns or is_warning_section
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = WARNING_HEADER_FILL if is_warning_header else HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in section["rows"]:
        sheet.append([stringify_export_value(section, row, index) for index in range(len(section["columns"]))])

    for row_index in range(4, sheet.max_row + 1):
        body_row_index = row_index - 4
        is_warning_row = body_row_index in warning_row_indexes
        for cell in sheet[row_index]:
            cell.font = Font(name="Segoe UI", size=10, color="243B53")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_warning_row:
                cell.fill = WARNING_BODY_FILL
            elif row_index % 2 == 0:
                cell.fill = BODY_ALT_FILL
        if is_warning_row:
            sheet.cell(row=row_index, column=1).font = Font(name="Segoe UI", size=10, bold=True, color=WARNING_TEXT_COLOR)

    sheet.freeze_panes = "A4"
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 22
    autosize_sheet(sheet)


def build_xlsx_bytes(exam: dict[str, Any], session: dict[str, Any]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sections = build_report_sections(exam, session, response_chunk_size=None)

    for section in sections:
        title = section["title"][:31] or "Rapor"
        sheet = workbook.create_sheet(title)
        write_section_to_sheet(sheet, section)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_ods_bytes(exam: dict[str, Any], session: dict[str, Any]) -> bytes:
    document = OpenDocumentSpreadsheet()
    sections = build_report_sections(exam, session, response_chunk_size=None)

    title_paragraph_style = Style(name="ReportTitleParagraph", family="paragraph")
    title_paragraph_style.addElement(TextProperties(fontfamily="Cambria", fontsize="16pt", fontweight="bold", color="#FFFFFF"))
    title_paragraph_style.addElement(ParagraphProperties(textalign="center"))
    document.styles.addElement(title_paragraph_style)

    section_paragraph_style = Style(name="SectionTitleParagraph", family="paragraph")
    section_paragraph_style.addElement(TextProperties(fontfamily="Cambria", fontsize="12pt", fontweight="bold", color="#12355B"))
    document.styles.addElement(section_paragraph_style)

    header_paragraph_style = Style(name="HeaderParagraph", family="paragraph")
    header_paragraph_style.addElement(TextProperties(fontfamily="Segoe UI", fontsize="10pt", fontweight="bold", color="#FFFFFF"))
    document.styles.addElement(header_paragraph_style)

    body_paragraph_style = Style(name="BodyParagraph", family="paragraph")
    body_paragraph_style.addElement(TextProperties(fontfamily="Segoe UI", fontsize="9pt", color="#243B53"))
    document.styles.addElement(body_paragraph_style)

    warning_body_paragraph_style = Style(name="WarningBodyParagraph", family="paragraph")
    warning_body_paragraph_style.addElement(TextProperties(fontfamily="Segoe UI", fontsize="9pt", fontweight="bold", color="#B42318"))
    document.styles.addElement(warning_body_paragraph_style)

    warning_header_paragraph_style = Style(name="WarningHeaderParagraph", family="paragraph")
    warning_header_paragraph_style.addElement(TextProperties(fontfamily="Segoe UI", fontsize="10pt", fontweight="bold", color="#FFFFFF"))
    document.styles.addElement(warning_header_paragraph_style)

    warning_section_paragraph_style = Style(name="WarningSectionTitleParagraph", family="paragraph")
    warning_section_paragraph_style.addElement(TextProperties(fontfamily="Cambria", fontsize="12pt", fontweight="bold", color="#B42318"))
    document.styles.addElement(warning_section_paragraph_style)

    title_cell_style = Style(name="ReportTitleCell", family="table-cell")
    title_cell_style.addElement(TableCellProperties(backgroundcolor="#12355B", padding="0.08in"))
    document.automaticstyles.addElement(title_cell_style)

    section_cell_style = Style(name="SectionTitleCell", family="table-cell")
    section_cell_style.addElement(TableCellProperties(backgroundcolor="#D9E8FF", padding="0.05in"))
    document.automaticstyles.addElement(section_cell_style)

    header_cell_style = Style(name="HeaderCell", family="table-cell")
    header_cell_style.addElement(TableCellProperties(backgroundcolor="#2F6690", padding="0.04in", border="0.5pt solid #D7DEE8"))
    document.automaticstyles.addElement(header_cell_style)

    warning_header_cell_style = Style(name="WarningHeaderCell", family="table-cell")
    warning_header_cell_style.addElement(TableCellProperties(backgroundcolor="#B42318", padding="0.04in", border="0.5pt solid #D7DEE8"))
    document.automaticstyles.addElement(warning_header_cell_style)

    body_cell_style = Style(name="BodyCell", family="table-cell")
    body_cell_style.addElement(TableCellProperties(padding="0.03in", border="0.5pt solid #D7DEE8"))
    document.automaticstyles.addElement(body_cell_style)

    alt_body_cell_style = Style(name="AltBodyCell", family="table-cell")
    alt_body_cell_style.addElement(TableCellProperties(backgroundcolor="#F8FAFC", padding="0.03in", border="0.5pt solid #D7DEE8"))
    document.automaticstyles.addElement(alt_body_cell_style)

    warning_body_cell_style = Style(name="WarningBodyCell", family="table-cell")
    warning_body_cell_style.addElement(TableCellProperties(backgroundcolor="#FEF3F2", padding="0.03in", border="0.5pt solid #D7DEE8"))
    document.automaticstyles.addElement(warning_body_cell_style)

    for section in sections:
        table = OdfTable()
        table.setAttribute("name", section["title"][:80] or "Rapor")
        column_count = max(len(section["columns"]), 1)
        warning_columns = set(section.get("warning_columns") or [])
        warning_row_indexes = get_warning_row_indexes(section)
        is_warning_section = section.get("tone") == "warning"
        for _ in range(column_count):
            table.addElement(TableColumn())

        title_row = TableRow()
        for index in range(column_count):
            cell = TableCell(valuetype="string", stylename=title_cell_style)
            cell.addElement(P(stylename=title_paragraph_style, text=REPORT_TITLE if index == 0 else ""))
            title_row.addElement(cell)
        table.addElement(title_row)

        section_row = TableRow()
        for index in range(column_count):
            cell = TableCell(valuetype="string", stylename=warning_body_cell_style if is_warning_section else section_cell_style)
            cell.addElement(P(stylename=warning_section_paragraph_style if is_warning_section else section_paragraph_style, text=section["title"] if index == 0 else ""))
            section_row.addElement(cell)
        table.addElement(section_row)

        header_row = TableRow()
        for value in section["columns"]:
            is_warning_header = stringify(value) in warning_columns or is_warning_section
            cell = TableCell(valuetype="string", stylename=warning_header_cell_style if is_warning_header else header_cell_style)
            cell.addElement(P(stylename=warning_header_paragraph_style if is_warning_header else header_paragraph_style, text=stringify(value)))
            header_row.addElement(cell)
        table.addElement(header_row)

        for row_index, row in enumerate(section["rows"], start=1):
            table_row = TableRow()
            is_warning_row = (row_index - 1) in warning_row_indexes
            row_style = warning_body_cell_style if is_warning_row else (alt_body_cell_style if row_index % 2 == 0 else body_cell_style)
            for column_index in range(len(section["columns"])):
                value = stringify_export_value(section, row, column_index)
                cell = TableCell(valuetype="string", stylename=row_style)
                paragraph_style = warning_body_paragraph_style if is_warning_row and column_index == 0 else body_paragraph_style
                cell.addElement(P(stylename=paragraph_style, text=value or ""))
                table_row.addElement(cell)
            table.addElement(table_row)

        document.spreadsheet.addElement(table)

    buffer = io.BytesIO()
    document.save(buffer)
    return buffer.getvalue()


def register_pdf_fonts() -> tuple[str, str]:
    try:
        pdfmetrics.getFont("SessionExportFont")
        pdfmetrics.getFont("SessionExportFontBold")
        return "SessionExportFont", "SessionExportFontBold"
    except KeyError:
        pass

    for font_paths in WINDOWS_FONT_CANDIDATES:
        if font_paths["regular"].exists() and font_paths["bold"].exists():
            pdfmetrics.registerFont(TTFont("SessionExportFont", str(font_paths["regular"])))
            pdfmetrics.registerFont(TTFont("SessionExportFontBold", str(font_paths["bold"])))
            return "SessionExportFont", "SessionExportFontBold"
    return "Helvetica", "Helvetica-Bold"


def build_pdf_paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    text = stringify(value).replace(" | ", "<br/>")
    safe_text = html.escape(text).replace("\n", "<br/>").replace("&lt;br/&gt;", "<br/>") or "-"
    return Paragraph(safe_text, style)


def compute_pdf_column_widths(section: dict[str, Any], available_width: float) -> list[float]:
    text_rows = [section["columns"], *[[stringify(value) for value in row] for row in section["rows"]]]
    estimated = []
    for index in range(len(section["columns"])):
        max_length = max(len(row[index]) if index < len(row) else 0 for row in text_rows)
        estimated.append(min(max(max_length, 8), 24))
    total = sum(estimated) or 1
    return [available_width * (value / total) for value in estimated]


def make_pdf_table(section: dict[str, Any], font_name: str, available_width: float) -> PdfTable:
    column_count = max(len(section["columns"]), 1)
    if column_count <= 6:
        header_font_size = 8
        body_font_size = 7
    elif column_count <= 12:
        header_font_size = 7
        body_font_size = 6.5
    elif column_count <= 18:
        header_font_size = 6.5
        body_font_size = 6
    else:
        header_font_size = 6
        body_font_size = 5.4

    header_style = ParagraphStyle(
        "SessionExportHeader",
        fontName=font_name,
        fontSize=header_font_size,
        leading=header_font_size + 1,
        textColor=colors.HexColor("#1F2937"),
        wordWrap="CJK",
    )
    warning_header_style = ParagraphStyle(
        "SessionExportHeaderWarning",
        parent=header_style,
        textColor=colors.white,
    )
    body_style = ParagraphStyle(
        "SessionExportBody",
        fontName=font_name,
        fontSize=body_font_size,
        leading=body_font_size + 1.5,
        wordWrap="CJK",
    )
    warning_label_style = ParagraphStyle(
        "SessionExportWarningLabel",
        parent=body_style,
        fontName=font_name,
        textColor=colors.HexColor("#B42318"),
    )
    warning_body_style = ParagraphStyle(
        "SessionExportWarningBody",
        parent=body_style,
        textColor=colors.HexColor("#7A271A"),
    )

    warning_columns = set(section.get("warning_columns") or [])
    warning_row_indexes = get_warning_row_indexes(section)
    is_warning_section = section.get("tone") == "warning"

    data = [
        [
            build_pdf_paragraph(value, warning_header_style if stringify(value) in warning_columns or is_warning_section else header_style)
            for value in section["columns"]
        ],
        *[
            [
                build_pdf_paragraph(
                    stringify_export_value(section, row, index),
                    warning_label_style if row_index in warning_row_indexes and index == 0 else (warning_body_style if row_index in warning_row_indexes else body_style),
                )
                for index in range(len(section["columns"]))
            ]
            for row_index, row in enumerate(section["rows"])
        ],
    ]
    table = PdfTable(data, repeatRows=1, colWidths=compute_pdf_column_widths(section, available_width), splitByRow=1)
    style_commands: list[tuple[Any, ...]] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E8FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#A3B1C6")),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]
    if is_warning_section:
        style_commands.append(("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B42318")))
        style_commands.append(("TEXTCOLOR", (0, 0), (-1, 0), colors.white))
    for column_index, column_name in enumerate(section["columns"]):
        if stringify(column_name) in warning_columns:
            style_commands.append(("BACKGROUND", (column_index, 0), (column_index, 0), colors.HexColor("#B42318")))
            style_commands.append(("TEXTCOLOR", (column_index, 0), (column_index, 0), colors.white))
    for row_index in warning_row_indexes:
        table_row_index = row_index + 1
        style_commands.append(("BACKGROUND", (0, table_row_index), (-1, table_row_index), colors.HexColor("#FEF3F2")))

    table.setStyle(TableStyle(style_commands))
    return table


def build_pdf_bytes(exam: dict[str, Any], session: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    regular_font_name, bold_font_name = register_pdf_fonts()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName=bold_font_name,
        fontSize=20,
        leading=24,
        alignment=1,
        textColor=colors.HexColor("#12355B"),
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName=regular_font_name,
        fontSize=11,
        leading=14,
        alignment=1,
        textColor=colors.HexColor("#334E68"),
    )
    section_title_style = ParagraphStyle(
        "ReportSectionTitle",
        parent=styles["Heading2"],
        fontName=bold_font_name,
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#12355B"),
    )
    warning_section_title_style = ParagraphStyle(
        "ReportSectionTitleWarning",
        parent=section_title_style,
        textColor=colors.HexColor("#B42318"),
    )

    story = [
        Paragraph(REPORT_TITLE, title_style),
        Spacer(1, 6),
        Paragraph(f"{stringify(exam.get('exam_code'))} - {stringify(exam.get('title'))}", subtitle_style),
        Spacer(1, 12),
    ]

    for section in build_report_sections(exam, session, response_chunk_size=18):
        story.append(Paragraph(section["title"], warning_section_title_style if section.get("tone") == "warning" else section_title_style))
        story.append(Spacer(1, 6))
        story.append(make_pdf_table(section, regular_font_name, doc.width))
        story.append(Spacer(1, 12))

    doc.build(story)
    return buffer.getvalue()


def build_zip_bytes(exam: dict[str, Any], session: dict[str, Any]) -> bytes:
    base_name = f"{exam.get('exam_code', 'sinav')}_{session.get('session_id', 'oturum')}"
    files = {
        f"{base_name}.csv": build_csv_bytes(exam, session),
        f"{base_name}.xlsx": build_xlsx_bytes(exam, session),
        f"{base_name}.ods": build_ods_bytes(exam, session),
        f"{base_name}.pdf": build_pdf_bytes(exam, session),
        f"{base_name}.txt": build_txt_bytes(exam, session),
        f"{base_name}.json": build_json_bytes(exam, session),
    }
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_name, payload in files.items():
            archive.writestr(file_name, payload)
    return buffer.getvalue()


def build_session_export(exam: dict[str, Any], session: dict[str, Any], export_format: str) -> tuple[bytes, str, str]:
    format_name = export_format.strip().lower()
    base_name = f"{exam.get('exam_code', 'sinav')}_{session.get('session_id', 'oturum')}"
    if format_name == "csv":
        return build_csv_bytes(exam, session), "text/csv; charset=utf-8", f"{base_name}.csv"
    if format_name == "txt":
        return build_txt_bytes(exam, session), "text/plain; charset=utf-8", f"{base_name}.txt"
    if format_name == "xlsx":
        return build_xlsx_bytes(exam, session), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{base_name}.xlsx"
    if format_name == "ods":
        return build_ods_bytes(exam, session), "application/vnd.oasis.opendocument.spreadsheet", f"{base_name}.ods"
    if format_name == "pdf":
        return build_pdf_bytes(exam, session), "application/pdf", f"{base_name}.pdf"
    if format_name == "json":
        return build_json_bytes(exam, session), "application/json; charset=utf-8", f"{base_name}.json"
    if format_name == "zip":
        return build_zip_bytes(exam, session), "application/zip", f"{base_name}_tum-formatlar.zip"
    raise ValueError("Desteklenmeyen export formati. Beklenen: csv, txt, xlsx, ods, pdf, json, zip.")